import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from src.app.enums import gender_types, disability_types, disability_states
from src.database.mysql import engine
from sqlmodel import Session, select
from src.models.inquiry import Inquiry
from src.models.user import User


def generate_report(date_range=None):
    with Session(engine) as session:
        inquiries = get_all_inquiries_with_user_data(session, date_range)
        stats = process_inquiry_statistics(inquiries)

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    t = now.strftime("%Y-%m-%d-%H-%M")
    directory = f"/app/user_data/{date}"
    file_path = f"{directory}/{t}-full-report.xlsx"
    os.makedirs(directory, exist_ok=True)
    export_statistics_to_excel(stats, filename=file_path)
    return file_path


def get_all_inquiries_with_user_data(session: Session, date_range=None):
    if date_range:
        start_date, end_date = date_range
        result = session.execute(
            select(User.gender, User.region, User.disability_type, User.disability_state, Inquiry.section_name)
            .join(Inquiry, User.user_id == Inquiry.user_id)
            .where(Inquiry.created_at >= start_date, Inquiry.created_at <= end_date)
        ).all()
    else:
        result = session.execute(
            select(User.gender, User.region, User.disability_type, User.disability_state, Inquiry.section_name)
            .join(Inquiry, User.user_id == Inquiry.user_id)
        ).all()

    return result


def process_inquiry_statistics(data):
    # Create a DataFrame from the query result
    df = pd.DataFrame(data, columns=["Jinsi", "Joyloashuv Hududi", "Nogironlik Guruhi", "Nogironlik Holati", "Bo'lim"])
    df["Jinsi"] = df["Jinsi"].map(gender_types)
    df["Nogironlik Guruhi"] = df["Nogironlik Guruhi"].map(disability_types)
    df["Nogironlik Holati"] = df["Nogironlik Holati"].map(disability_states)
    df["Joyloashuv Hududi"] = df["Joyloashuv Hududi"].apply(lambda x: x.value)
    # Create pivot tables for each category
    gender_pivot = df.pivot_table(index="Jinsi", columns="Bo'lim", aggfunc="size", fill_value=0)
    region_pivot = df.pivot_table(index="Joyloashuv Hududi", columns="Bo'lim", aggfunc="size", fill_value=0)
    disability_type_pivot = df.pivot_table(index="Nogironlik Guruhi", columns="Bo'lim", aggfunc="size", fill_value=0)
    disability_state_pivot = df.pivot_table(index="Nogironlik Holati", columns="Bo'lim", aggfunc="size", fill_value=0)

    # Add a 'Total' column to each pivot table
    for pivot in [gender_pivot, region_pivot, disability_type_pivot, disability_state_pivot]:
        pivot["Jami"] = pivot.sum(axis=1)

    # Add a 'Total' row to each pivot table
    gender_pivot.loc["Jami"] = gender_pivot.sum(axis=0)
    region_pivot.loc["Jami"] = region_pivot.sum(axis=0)
    disability_type_pivot.loc["Jami"] = disability_type_pivot.sum(axis=0)
    disability_state_pivot.loc["Jami"] = disability_state_pivot.sum(axis=0)

    # return gender_pivot, region_pivot, disability_type_pivot, disability_state_pivot
    return {
        "Jinsi Bo'yicha": gender_pivot,
        "Joyloashuv Hududi bo'yicha": region_pivot,
        "Nogironlik Guruhi Bo'yicha": disability_type_pivot,
        "Nogironlik Holati Bo'yicha": disability_state_pivot
    }


def export_statistics_to_excel(statistics, filename="/app/user_data/inquiry_statistics.xlsx", spacing=2):
    # Create a new workbook and a sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Murojaatlar statistikasi"

    current_row = 1

    for table_name, table_data in statistics.items():
        # Write the table name as a header
        ws.cell(row=current_row, column=1, value=table_name)

        # Convert the DataFrame to rows and write to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(table_data.reset_index(), index=False, header=True), start=current_row + 1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Move to the next position after spacing
        current_row += len(table_data) + spacing + 2  # +2 for header and totals row

    # Save the workbook
    wb.save(filename)
